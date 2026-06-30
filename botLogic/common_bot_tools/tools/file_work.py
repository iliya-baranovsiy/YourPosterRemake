import asyncio
import os
import time
from pathlib import Path
from openpyxl import load_workbook
from aiogram.types import Message


class FileWork:
    def __init__(self, message: Message):
        self.message = message
        self.path = Path("botLogic") / "src" / "user_files" / f"{self.message.chat.id}.xlsx"

    def _parse_file(self, channel_id: int):
        wb = load_workbook(self.path, read_only=True)
        ws = wb.active
        result = []
        for a, b in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if a is None and b is None:
                continue
            if len(str(a)) >= 50 or len(str(b)) >= 3600:
                continue
            result.append({'channel_id': channel_id, 'title': str(a), "content": str(b)})
        wb.close()
        os.remove(path=self.path)
        return result if result else None

    async def handle_exel(self, channel_id: int):
        await self._save_file()
        return await asyncio.to_thread(self._parse_file, channel_id)

    async def _save_file(self):
        await self.message.bot.download(
            file=self.message.document,
            destination=self.path
        )
